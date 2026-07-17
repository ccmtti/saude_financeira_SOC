import json
import ast
import pandas as pd
import calendar
import io
import base64
import logging
import threading
import concurrent.futures
from datetime import datetime
import time
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import streamlit as st

# Desabilita logs longos e desnecessários da biblioteca na tela
logging.getLogger('zeep').setLevel(logging.ERROR)

try:
    from zeep import Client
except ImportError:
    st.error(
        "**Biblioteca `zeep` não encontrada.** "
        "Execute no terminal: `pip install zeep`"
    )
    st.stop()

# ==========================================
# CONFIGURAÇÃO DA PÁGINA STREAMLIT
# ==========================================
st.set_page_config(
    page_title="Saúde Financeira do Contrato — SOC",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ==========================================
# CSS CUSTOMIZADO — visual premium
# ==========================================
st.markdown("""
<style>
    /* ---------- Google Font ---------- */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
    }

    /* ---------- Sidebar ---------- */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #005850 0%, #007b6b 100%);
    }
    section[data-testid="stSidebar"] * {
        color: #e8f5f3 !important;
    }
    section[data-testid="stSidebar"] .stNumberInput label,
    section[data-testid="stSidebar"] .stTextInput label {
        font-weight: 600;
        letter-spacing: 0.02em;
    }

    /* ---------- Metric cards ---------- */
    div[data-testid="stMetric"] {
        background: linear-gradient(135deg, #005850 0%, #007b6b 100%);
        border: 1px solid #009f8f;
        border-radius: 12px;
        padding: 14px 16px;
        box-shadow: 0 4px 14px rgba(0,88,80,.3);
        transition: transform .2s ease, box-shadow .2s ease;
        overflow: hidden;
    }
    div[data-testid="stMetric"]:hover {
        transform: translateY(-3px);
        box-shadow: 0 8px 24px rgba(0,88,80,.45);
    }
    div[data-testid="stMetric"] label {
        color: #a7dbd6 !important;
        font-size: .72rem !important;
        text-transform: uppercase;
        letter-spacing: .04em;
        white-space: nowrap;
    }
    div[data-testid="stMetric"] [data-testid="stMetricValue"] {
        color: #ffffff !important;
        font-weight: 700;
        font-size: .95rem !important;
        word-break: break-word;
        white-space: normal !important;
        line-height: 1.3;
    }

    /* ---------- Section titles ---------- */
    .section-title {
        font-size: 1.25rem;
        font-weight: 700;
        color: #f1f5f9;
        border-left: 4px solid #009f8f;
        padding-left: 14px;
        margin: 2rem 0 1rem;
    }

    /* ---------- Hero banner ---------- */
    .hero-banner {
        background: linear-gradient(135deg, #007b6b 0%, #005850 50%, #003d35 100%);
        border-radius: 16px;
        padding: 2.5rem 2rem;
        margin-bottom: 2rem;
        border: 1px solid #009f8f;
        box-shadow: 0 8px 32px rgba(0,88,80,.35);
        text-align: center;
    }
    .hero-banner h1 {
        font-size: 2rem;
        font-weight: 800;
        background: linear-gradient(90deg, #4dd9c0, #ffffff, #80e8d6);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: .3rem;
    }
    .hero-banner p {
        color: #a7dbd6;
        font-size: .95rem;
        margin: 0;
    }

    /* ---------- DataFrames ---------- */
    .stDataFrame {
        border-radius: 10px;
        overflow: hidden;
    }

    /* ---------- Tabela só para impressão ---------- */
    .print-only-table, .print-only-status {
        display: none !important;
    }

    /* ---------- Download button ---------- */
    .stDownloadButton > button {
        background: linear-gradient(135deg, #007b6b, #009f8f) !important;
        color: white !important;
        border: none !important;
        border-radius: 10px !important;
        padding: 0.65rem 2rem !important;
        font-weight: 600 !important;
        letter-spacing: .02em;
        transition: all .2s ease !important;
    }
    .stDownloadButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(0,159,143,.4) !important;
    }

    /* ---------- Info / Warning boxes ---------- */
    .stAlert {
        border-radius: 10px !important;
    }

    /* ---------- Sidebar button ---------- */
    section[data-testid="stSidebar"] .stButton > button {
        width: 100%;
        background: linear-gradient(135deg, #009f8f, #4dd9c0) !important;
        color: #003d35 !important;
        border: none !important;
        border-radius: 10px !important;
        padding: .7rem 1rem !important;
        font-weight: 700 !important;
        font-size: .95rem !important;
        letter-spacing: .03em;
        margin-top: .5rem;
        transition: all .25s ease !important;
    }
    section[data-testid="stSidebar"] .stButton > button:hover {
        box-shadow: 0 6px 22px rgba(0,159,143,.5) !important;
        transform: translateY(-2px);
    }

    /* ========================================
       IMPRESSÃO — A4 Paisagem
       ======================================== */
    @media print {
        @page {
            size: A4 portrait;
            margin: 4mm 10mm;
        }

        /* Esconde elementos não imprimíveis */
        section[data-testid="stSidebar"],
        .stStatusWidget,
        .stDeployButton,
        header[data-testid="stHeader"],
        .hero-banner,
        .stDownloadButton,
        .print-btn,
        .stButton,
        footer,
        #MainMenu,
        [data-testid="stToolbar"],
        [data-testid="stStatusWidget"],
        [data-testid="stExpander"],
        iframe,
        hr { display: none !important; }

        /* Esconde o st.dataframe (canvas) — usamos tabela HTML no lugar */
        .stDataFrame,
        [data-testid="stDataFrame"] {
            display: none !important;
        }

        /* MOSTRA a tabela HTML e o status só na impressão */
        .print-only-table {
            display: block !important;
        }
        .print-only-status {
            display: block !important;
            font-size: 0.85rem;
            color: #005850;
            font-weight: 600;
            margin-bottom: 1rem;
            border-bottom: 1px solid #b0d4cf;
            padding-bottom: 5px;
        }

        /* Reset backgrounds para branco e ativa zoom no Chrome */
        html, body,
        .main,
        .block-container,
        section[data-testid="stMain"],
        [data-testid="stAppViewContainer"],
        [data-testid="stAppViewBlockContainer"] {
            background: white !important;
            color: #111 !important;
            padding: 0 !important;
            margin: 0 !important;
            overflow: visible !important;
            height: auto !important;
        }

        /* Reset total da escala e largura para o padrão natural */
        body, html {
            zoom: 100% !important;
            width: 100% !important;
            max-width: 100% !important;
            margin: 0 !important;
            padding: 0 !important;
            overflow-x: hidden !important;
        }

        /* Fallback específico para o Mozilla Firefox */
        @-moz-document url-prefix() {
            body {
                transform: none !important;
                width: 100% !important;
            }
        }

        /* BLINDAGEM MÁXIMA DA LARGURA */
        .block-container,
        section[data-testid="stMain"],
        [data-testid="stAppViewContainer"],
        [data-testid="stAppViewBlockContainer"],
        [data-testid="stMarkdownContainer"],
        .print-only-table {
            max-width: 100vw !important;
            min-width: 100% !important;
            width: 100% !important;
            padding: 0 !important;
            box-sizing: border-box !important;
        }
        
        /* Permite que colunas possam se comportar naturalmente */
        [data-testid="stHorizontalBlock"] {
            display: flex !important;
            flex-wrap: wrap !important;
            width: 100% !important;
        }

        /* Metric cards para impressão */
        div[data-testid="stMetric"] {
            background: #f0faf8 !important;
            border: 1px solid #009f8f !important;
            box-shadow: none !important;
            break-inside: avoid;
            padding: 6px 8px !important;
        }
        div[data-testid="stMetric"] label {
            color: #005850 !important;
            font-size: .6rem !important;
        }
        div[data-testid="stMetric"] [data-testid="stMetricValue"] {
            color: #003d35 !important;
            font-size: .78rem !important;
        }

        /* Títulos de seção */
        .section-title {
            color: #005850 !important;
            border-left-color: #009f8f !important;
            font-size: .95rem !important;
            margin: .8rem 0 .4rem !important;
            break-after: avoid;
        }

        /* Estilo da tabela HTML de impressão */
        .print-only-table table {
            width: 100% !important;
            max-width: 100% !important;
            border-collapse: collapse;
            font-size: 9pt !important; /* Tamanho confortável para leitura em papel */
            color: #111;
            margin-bottom: .8rem;
            table-layout: fixed !important; /* O SEGREDO para não deixar cortar a direita */
        }
        .print-only-table th {
            background: #f0faf8;
            color: #005850;
            font-weight: 600;
            border: 1px solid #b0d4cf;
            padding: 4px;
            text-align: left;
            word-wrap: break-word !important;
            white-space: normal !important;
            overflow-wrap: break-word !important;
        }
        .print-only-table td {
            background: white;
            color: #222;
            border: 1px solid #cbd5e1;
            padding: 4px;
            word-wrap: break-word !important;
            white-space: normal !important;
            overflow-wrap: break-word !important;
        }
        .print-only-table tr:nth-child(even) td {
            background: #f8fdfb;
        }

        /* Info de avulsos vazio */
        .print-info {
            display: block !important;
            font-size: .8rem;
            color: #475569;
            font-style: italic;
            margin-top: .3rem;
        }

        /* Evita quebras dentro de colunas */
        [data-testid="stHorizontalBlock"] {
            break-inside: avoid;
        }
    }
</style>
""", unsafe_allow_html=True)

# ==========================================
# SISTEMA DE LOGIN DE SEGURANÇA
# ==========================================
if "autenticado" not in st.session_state:
    st.session_state["autenticado"] = False

if not st.session_state["autenticado"]:
    st.markdown('<div class="hero-banner" style="margin-top: 3rem;"><h1>🔒 Acesso Restrito</h1><p>Insira suas credenciais corporativas para acessar o painel.</p></div>', unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        with st.form("login_form"):
            usuario = st.text_input("Usuário")
            senha = st.text_input("Senha", type="password")
            submit = st.form_submit_button("Entrar no Sistema")
            
            if submit:
                # Agora o sistema puxa a senha dos "Secrets" do Streamlit (ou usa um padrão caso rode localmente)
                try:
                    senha_correta = st.secrets.get("SENHA_SISTEMA", "consulta@04109")
                    usuario_correto = st.secrets.get("USUARIO_SISTEMA", "ccmt")
                except Exception:
                    senha_correta = "consulta@04109"
                    usuario_correto = "ccmt"

                if usuario == usuario_correto and senha == senha_correta:
                    st.session_state["autenticado"] = True
                    st.rerun()
                else:
                    st.error("Usuário ou senha incorretos. Acesso negado.")
    st.stop()

# ==========================================
# CONFIGURAÇÕES DE ACESSO AO SOC
# ==========================================
EMPRESA_LOGADA = "323506"
WSDL_SOC_SOAP = "https://ws1.soc.com.br/WSSoc/services/ExportaDadosWs?wsdl"

# Para deixar o GitHub "Público" de forma segura, as chaves do SOC devem ir para o "Secrets" do Streamlit!
def get_soc_auth(nome_secret_cod, nome_secret_chave, default_cod, default_chave):
    try:
        return {"codigo": st.secrets[nome_secret_cod], "chave": st.secrets[nome_secret_chave]}
    except Exception:
        # Fallback de uso local (Não suba essas chaves reais num repositório público se possível)
        return {"codigo": default_cod, "chave": default_chave}

EXPORTA_FATURAMENTO = get_soc_auth("SOC_COD_FAT", "SOC_CHAVE_FAT", "187006", "fb996f2d1a16d2629600")
EXPORTA_FUNCIONARIOS = get_soc_auth("SOC_COD_FUNC", "SOC_CHAVE_FUNC", "217724", "d9bdc2c1010858b03175")
EXPORTA_PRECOS = get_soc_auth("SOC_COD_PRECO", "SOC_CHAVE_PRECO", "195953", "131d6afed38b86e21a03")

# Trava para organizar as mensagens quando usar Multithreading
print_lock = threading.Lock()


# ==========================================
# CONEXÃO COM O WEBSERVICE (cache global)
# ==========================================
@st.cache_resource(show_spinner=False)
def conectar_soc():
    """Abre a conexão com o WebService uma única vez e reutiliza."""
    return Client(WSDL_SOC_SOAP)


# ==========================================
# FUNÇÕES DE NEGÓCIO — preservadas intactas
# ==========================================
def fazer_requisicao_soc(cliente_soap, payload, nome_rotina, log_container=None):
    """Acessa o SOC diretamente via WebService (SOAP), ignorando a rota REST."""
    with print_lock:
        if log_container is not None:
            log_container.write(f"🔄 `[{nome_rotina}]` → Consultando...")

    try:
        parametro_json = json.dumps(payload, separators=(',', ':'))

        vo = {
            'parametros': parametro_json,
            'erro': False
        }

        max_retries = 3
        retry_delay = 2
        resultado = None

        for tentativa in range(1, max_retries + 1):
            try:
                resultado = cliente_soap.service.exportaDadosWs(arg0=vo)
                break
            except Exception as e:
                if tentativa < max_retries:
                    with print_lock:
                        if log_container is not None:
                            log_container.write(f"⚠️ `[{nome_rotina}]` → Falha de rede (tentativa {tentativa}/{max_retries}). Retentando...")
                    time.sleep(retry_delay)
                else:
                    raise e

        if not resultado:
            return []

        resultado_str = ""
        if hasattr(resultado, 'retorno') and resultado.retorno:
            resultado_str = str(resultado.retorno).strip()
        elif hasattr(resultado, 'arquivo') and resultado.arquivo:
            resultado_str = base64.b64decode(resultado.arquivo).decode('utf-8', errors='replace').strip()
        else:
            resultado_str = str(resultado).strip()

        if not resultado_str:
            with print_lock:
                if log_container is not None:
                    log_container.write(f"⚠️ `{nome_rotina}`: Nenhum dado retornado para o período.")
            return []

        if resultado_str.startswith('[') or resultado_str.startswith('{'):
            # Tenta JSON puro primeiro; se falhar (aspas simples do zeep/str()),
            # usa ast.literal_eval como fallback seguro.
            try:
                dados = json.loads(resultado_str)
            except json.JSONDecodeError:
                try:
                    dados = ast.literal_eval(resultado_str)
                except (ValueError, SyntaxError):
                    with print_lock:
                        if log_container is not None:
                            log_container.error(
                                f"❌ `{nome_rotina}`: Resposta não é JSON nem Python válido.\n"
                                f"Primeiros 300 chars: `{resultado_str[:300]}`"
                            )
                    return []
            with print_lock:
                if log_container is not None:
                    log_container.write(
                        f"✅ `{nome_rotina}`: Sucesso! "
                        f"**{len(dados) if isinstance(dados, list) else 1}** registros."
                    )
            return dados
        else:
            df = pd.read_csv(io.StringIO(resultado_str), sep=';', dtype=str)
            df.columns = df.columns.str.strip()
            if df.empty:
                with print_lock:
                    if log_container is not None:
                        log_container.write(f"⚠️ `{nome_rotina}`: Consulta ok, mas sem dados no período.")
                return []
            dados = df.to_dict('records')
            with print_lock:
                if log_container is not None:
                    log_container.write(f"✅ `{nome_rotina}`: Sucesso! **{len(dados)}** registros.")
            return dados

    except Exception as e:
        with print_lock:
            if log_container is not None:
                log_container.error(f"❌ `{nome_rotina}`: {e}")
        return []


def buscar_funcionarios_ativos(cliente_soap, empresa_trabalho, log_container=None):
    payload = {
        "empresa": EMPRESA_LOGADA,
        "codigo": EXPORTA_FUNCIONARIOS["codigo"],
        "chave": EXPORTA_FUNCIONARIOS["chave"],
        "tipoSaida": "json",
        "empresaTrabalho": str(empresa_trabalho),
        "cpf": "",
        "situacaoFuncionario": "S,A",
        "parametroData": "0",
        "dataInicio": "",
        "dataFim": ""
    }
    dados = fazer_requisicao_soc(cliente_soap, payload, "Busca de Funcionários (Ativos/Afastados)", log_container)
    return len(dados) if isinstance(dados, list) else 0


def buscar_faturamento_historico(cliente_soap, empresa_trabalho, meses_historico, log_container=None):
    """Busca os dados de todos os meses simultaneamente respeitando o limite do SOC"""
    dados_faturamento = []
    data_atual = datetime.now()
    # Coleta os rótulos de cada mês para log posterior (thread-safe)
    meses_labels = []

    def buscar_mes(i):
        data_alvo = data_atual - relativedelta(months=i)
        mes = str(data_alvo.month).zfill(2)
        ano = str(data_alvo.year)
        ultimo_dia = str(calendar.monthrange(int(ano), int(mes))[1]).zfill(2)

        payload = {
            "empresa": EMPRESA_LOGADA,
            "codigo": EXPORTA_FATURAMENTO["codigo"],
            "chave": EXPORTA_FATURAMENTO["chave"],
            "tipoSaida": "json",
            "diaInicioCobranca": "01",
            "diaFimCobranca": ultimo_dia,
            "mes": mes,
            "ano": ano,
            "empresaTrabalho": str(empresa_trabalho),
            "unidade": ""
        }

        rotina = f"Faturamento ({mes}/{ano})"
        # IMPORTANTE: NÃO passar log_container para threads secundárias —
        # Streamlit levanta NoSessionContext ao chamar st.write() fora da main thread.
        dados = fazer_requisicao_soc(cliente_soap, payload, rotina, None)
        
        # Injeta a ordem cronológica no dicionário para podermos pegar o valor mais atual depois
        if isinstance(dados, list):
            for d in dados:
                d['_ANO_MES_ORDEM'] = f"{ano}-{mes}"
                
        return dados

    # LIMITE DE SEGURANÇA: max_workers=4
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        futures = [executor.submit(buscar_mes, i) for i in range(meses_historico)]
        for future in concurrent.futures.as_completed(futures):
            retorno = future.result()
            if isinstance(retorno, list):
                dados_faturamento.extend(retorno)

    # Log resumido na thread principal (seguro para Streamlit)
    if log_container is not None:
        log_container.write(
            f"✅ Faturamento: **{len(dados_faturamento)}** registros coletados "
            f"em **{meses_historico}** meses."
        )

    return pd.DataFrame(dados_faturamento)


def buscar_precos_contrato(cliente_soap, empresa_trabalho, log_container=None):
    payload = {
        "empresa": EMPRESA_LOGADA,
        "codigo": EXPORTA_PRECOS["codigo"],
        "chave": EXPORTA_PRECOS["chave"],
        "tipoSaida": "json",
        "codigoEmpresa": str(empresa_trabalho),
        "codigoUnidade": "",
        "codigoProduto": "",
        "codigoGrupoProduto": ""
    }
    dados = fazer_requisicao_soc(cliente_soap, payload, "Regras de Contrato", log_container)
    return pd.DataFrame(dados) if isinstance(dados, list) else pd.DataFrame()


def tratar_valor_numerico(df, coluna):
    if coluna in df.columns:
        return pd.to_numeric(df[coluna].astype(str).str.replace(',', '.'), errors='coerce').fillna(0)
    return 0


def processar_dados(df_fat, df_precos, qtd_ativos, meses_historico):
    """
    Processa DataFrames e retorna os artefatos necessários para o dashboard
    e para a geração da planilha Excel.  Lógica de Pandas 100 % preservada.
    """
    if df_fat.empty:
        return None

    df_fat = df_fat.copy()
    df_precos = df_precos.copy()

    df_fat.columns = df_fat.columns.str.upper()
    
    # 🚫 REGRA NOVA: Excluir totalmente produtos de DESCONTO da análise
    if 'NOME_PRODUTO' in df_fat.columns:
        mask_desconto = df_fat['NOME_PRODUTO'].astype(str).str.upper().str.contains('DESCONTO', na=False)
        df_fat = df_fat[~mask_desconto].copy()

    if df_fat.empty:
        return None
    if not df_precos.empty:
        df_precos.columns = [col.strip() for col in df_precos.columns]

    for col in ['VALOR_TOTAL_PRODUTO', 'QUANTIDADE_VIDAS_ATIVAS']:
        df_fat[col] = tratar_valor_numerico(df_fat, col)

    razao_social = df_fat['RAZAO_SOCIAL'].iloc[0] if 'RAZAO_SOCIAL' in df_fat.columns else "Não Identificada"

    data_assinatura = "Não encontrada"
    col_assinatura = 'dataAssinaturaContrato' if 'dataAssinaturaContrato' in df_precos.columns else 'DATAASSINATURACONTRATO'

    if not df_precos.empty and col_assinatura in df_precos.columns:
        datas_validas = df_precos[col_assinatura].dropna().unique()
        datas_validas = [d for d in datas_validas if str(d).strip()]
        if datas_validas:
            data_assinatura = datas_validas[0]

    # Ordena para garantir que o "last()" no groupby pegue o mês mais recente
    if '_ANO_MES_ORDEM' in df_fat.columns:
        df_fat = df_fat.sort_values(by='_ANO_MES_ORDEM', ascending=True)

    df_fat['Custo_Por_Funcionario'] = df_fat.apply(
        lambda row: row['VALOR_TOTAL_PRODUTO'] / row['QUANTIDADE_VIDAS_ATIVAS']
        if row['QUANTIDADE_VIDAS_ATIVAS'] > 0 else 0,
        axis=1
    )

    resumo_produtos = df_fat.groupby(['CODIGO_PRODUTO', 'NOME_PRODUTO']).agg(
        Total_Cobrado_Periodo=('VALOR_TOTAL_PRODUTO', 'sum'),
        Ultima_Qtd_Vidas=('QUANTIDADE_VIDAS_ATIVAS', 'last'),
        Custo_Medio_Por_Vida=('Custo_Por_Funcionario', 'last')
    ).reset_index()

    if not df_precos.empty:
        # Helper para encontrar a coluna independente de ser Maiúscula ou Minúscula
        def get_col(df, nome_alvo):
            for c in df.columns:
                if str(c).lower() == nome_alvo.lower():
                    return c
            return nome_alvo

        col_prod = get_col(df_precos, 'codigoProduto')
        col_vmin = get_col(df_precos, 'valorMinimo')
        col_mvid = get_col(df_precos, 'minimoVidas')
        col_vvm = get_col(df_precos, 'valorVidaMes')
        col_vm = get_col(df_precos, 'valorMensal')

        # Garantir que extraímos as colunas que existem
        cols_extract = [col_prod, col_vmin, col_mvid]
        if col_vvm in df_precos.columns: cols_extract.append(col_vvm)
        if col_vm in df_precos.columns: cols_extract.append(col_vm)

        df_precos_red = df_precos[cols_extract].copy()
        df_precos_red.rename(columns={col_prod: 'CODIGO_PRODUTO'}, inplace=True)

        df_precos_red[col_vmin] = tratar_valor_numerico(df_precos_red, col_vmin)
        df_precos_red[col_mvid] = tratar_valor_numerico(df_precos_red, col_mvid)
        
        if col_vvm in df_precos_red.columns:
            df_precos_red[col_vvm] = tratar_valor_numerico(df_precos_red, col_vvm)
        if col_vm in df_precos_red.columns:
            df_precos_red[col_vm] = tratar_valor_numerico(df_precos_red, col_vm)
            
        def classificar_cobranca(row):
            # Se for > 0 em valorVidaMes ou valorMensal, o SOC considera cobrança contínua (Mensalidade)
            v = row.get(col_vvm, 0)
            m = row.get(col_vm, 0)
            if v > 0 or m > 0:
                return True
            return False

        df_precos_red['Is_Mensalidade_SOC'] = df_precos_red.apply(classificar_cobranca, axis=1)
        df_precos_red.drop_duplicates(subset=['CODIGO_PRODUTO'], inplace=True)

        # Guarda a lista de códigos que são mensalidades segundo o SOC
        codigos_mensalidades = df_precos_red[df_precos_red['Is_Mensalidade_SOC']]['CODIGO_PRODUTO'].tolist()

        resumo_produtos = pd.merge(resumo_produtos, df_precos_red, on='CODIGO_PRODUTO', how='left')
        resumo_produtos.rename(columns={
            col_vmin: 'Faturamento_Minimo_Valor (R$)',
            col_mvid: 'Faturamento_Minimo_Vidas',
            'Ultima_Qtd_Vidas': 'Media_Vidas_Cobradas' # Mantém o nome interno igual para não quebrar a UI
        }, inplace=True)
    else:
        codigos_mensalidades = []
        resumo_produtos['Faturamento_Minimo_Valor (R$)'] = 0
        resumo_produtos['Faturamento_Minimo_Vidas'] = 0
        resumo_produtos.rename(columns={'Ultima_Qtd_Vidas': 'Media_Vidas_Cobradas'}, inplace=True)

    # Identifica se é mensalidade pelo NOME ou pela REGRA DE PREÇO (SOC)
    nome_prod_resumo_upper = resumo_produtos['NOME_PRODUTO'].astype(str).str.upper()
    mask_mens = (
        nome_prod_resumo_upper.str.startswith('MENSALIDADE', na=False) |
        nome_prod_resumo_upper.str.startswith('MENSAGERIA', na=False) |
        nome_prod_resumo_upper.str.startswith('ENVIO DIRETO', na=False) |
        resumo_produtos['CODIGO_PRODUTO'].isin(codigos_mensalidades)
    )

    # Separação Original (Excel)
    resumo_produtos_mensalidades = resumo_produtos[mask_mens].copy()
    resumo_produtos_demais = resumo_produtos[~mask_mens].copy()

    # Separação Sintese (Tela)
    resumo_produtos_sintese = resumo_produtos.drop(columns=['CODIGO_PRODUTO', 'Is_Mensalidade_SOC', col_vvm, col_vm], errors='ignore')
    resumo_mensalidades = resumo_produtos_sintese[mask_mens].copy()
    resumo_demais = resumo_produtos_sintese[~mask_mens].copy()

    # Adicionar Totalizadores de Mensalidades
    if not resumo_mensalidades.empty:
        soma_custo_vida = resumo_mensalidades.loc[resumo_mensalidades['Media_Vidas_Cobradas'] > 0, 'Custo_Medio_Por_Vida'].sum()
        linha_total_sintese = pd.DataFrame([{
            'NOME_PRODUTO': '➡️ TOTAL',
            'Total_Cobrado_Periodo': resumo_mensalidades['Total_Cobrado_Periodo'].sum(),
            'Media_Vidas_Cobradas': None,
            'Custo_Medio_Por_Vida': soma_custo_vida,
            'Faturamento_Minimo_Valor (R$)': resumo_mensalidades['Faturamento_Minimo_Valor (R$)'].sum(),
            'Faturamento_Minimo_Vidas': None
        }])
        resumo_mensalidades = pd.concat([resumo_mensalidades, linha_total_sintese], ignore_index=True)

    if not resumo_produtos_mensalidades.empty:
        soma_custo_vida = resumo_produtos_mensalidades.loc[resumo_produtos_mensalidades['Media_Vidas_Cobradas'] > 0, 'Custo_Medio_Por_Vida'].sum()
        linha_total_excel = pd.DataFrame([{
            'CODIGO_PRODUTO': '',
            'NOME_PRODUTO': '➡️ TOTAL',
            'Total_Cobrado_Periodo': resumo_produtos_mensalidades['Total_Cobrado_Periodo'].sum(),
            'Media_Vidas_Cobradas': None,
            'Custo_Medio_Por_Vida': soma_custo_vida,
            'Faturamento_Minimo_Valor (R$)': resumo_produtos_mensalidades['Faturamento_Minimo_Valor (R$)'].sum(),
            'Faturamento_Minimo_Vidas': None
        }])
        resumo_produtos_mensalidades = pd.concat([resumo_produtos_mensalidades, linha_total_excel], ignore_index=True)

    # Filtros de produtos Base Completa (df_fat)
    nome_produto_upper = df_fat['NOME_PRODUTO'].astype(str).str.upper()
    
    mask_mensalidade_df = (
        nome_produto_upper.str.startswith('MENSALIDADE', na=False) |
        nome_produto_upper.str.startswith('MENSAGERIA', na=False) |
        nome_produto_upper.str.startswith('ENVIO DIRETO', na=False) |
        df_fat['CODIGO_PRODUTO'].isin(codigos_mensalidades)
    )
    mask_exames = nome_produto_upper.str.startswith('EXAME', na=False)
    mask_mensageria = nome_produto_upper.str.startswith('MENSAGERIA', na=False)

    df_mensalidades = df_fat[mask_mensalidade_df].copy()

    mask_avulsos = ~(mask_mensalidade_df | mask_exames | mask_mensageria)
    df_avulsos = df_fat[mask_avulsos].copy()

    avulsos_detalhado = df_avulsos[
        ['DATA_COBRANCA', 'NOME_PRODUTO', 'VALOR_TOTAL_PRODUTO']
    ].copy()
    if 'DATA_COBRANCA' in avulsos_detalhado.columns:
        avulsos_detalhado.sort_values(by='DATA_COBRANCA', ascending=False, inplace=True)

    custo_total_mensalidade_por_funcionario = 0
    if 'Custo_Medio_Por_Vida' in resumo_mensalidades.columns:
        # Exclui a linha de "TOTAL" para não duplicar, caso exista
        df_calc = resumo_mensalidades[resumo_mensalidades['NOME_PRODUTO'] != '➡️ TOTAL']
        custo_total_mensalidade_por_funcionario = df_calc.loc[df_calc['Media_Vidas_Cobradas'] > 0, 'Custo_Medio_Por_Vida'].sum()

    data_atual = datetime.now()
    data_fim_str = data_atual.strftime('%m/%Y')
    data_inicio_str = (data_atual - relativedelta(months=meses_historico - 1)).strftime('%m/%Y')

    df_resumo_geral = pd.DataFrame([{
        "Razão Social": razao_social,
        "Data da Consulta": data_atual.strftime('%d/%m/%Y'),
        "Período Avaliado": f"{data_inicio_str} a {data_fim_str}",
        "Meses Analisados": meses_historico,
        "Data de Assinatura do Contrato": data_assinatura,
        "Funcionários Ativos Hoje (SOC)": qtd_ativos,
        "Soma das Mensalidades por Funcionário (R$)": round(custo_total_mensalidade_por_funcionario, 2),
        "Faturamento Total no Período (R$)": round(df_fat['VALOR_TOTAL_PRODUTO'].sum(), 2),
    }])

    return {
        "razao_social": razao_social,
        "df_resumo_geral": df_resumo_geral,
        "resumo_produtos": resumo_produtos, # Adicionado para corrigir a exportação multi-empresas
        "resumo_mensalidades": resumo_mensalidades,
        "resumo_demais": resumo_demais,
        "resumo_produtos_mensalidades": resumo_produtos_mensalidades,
        "resumo_produtos_demais": resumo_produtos_demais,
        "avulsos_detalhado": avulsos_detalhado,
        "df_fat": df_fat,
        "data_assinatura": data_assinatura,
        "custo_mensalidade_func": round(custo_total_mensalidade_por_funcionario, 2),
        "faturamento_total": round(df_fat['VALOR_TOTAL_PRODUTO'].sum(), 2),
        "qtd_ativos": qtd_ativos,
        "periodo": f"{data_inicio_str} a {data_fim_str}",
        "meses": meses_historico,
    }


def gerar_excel_em_memoria(resultado):
    """Gera a planilha Excel em memória (BytesIO) — mesma formatação original."""
    buf = io.BytesIO()

    df_resumo_geral = resultado["df_resumo_geral"]
    resumo_mensalidades = resultado["resumo_mensalidades"]
    resumo_demais = resultado["resumo_demais"]
    resumo_produtos_mensalidades = resultado["resumo_produtos_mensalidades"]
    resumo_produtos_demais = resultado["resumo_produtos_demais"]
    avulsos_detalhado = resultado["avulsos_detalhado"]
    df_fat = resultado["df_fat"]

    with pd.ExcelWriter(buf, engine='openpyxl') as writer:

        # 1. ABA SÍNTESE IMPRESSÃO
        aba_sintese = 'Síntese Impressão'

        linha_resumo = 2
        linha_mensal = linha_resumo + len(df_resumo_geral) + 4
        linha_demais = linha_mensal + len(resumo_mensalidades) + 4
        linha_avulsos = linha_demais + len(resumo_demais) + 4

        df_resumo_geral.to_excel(writer, sheet_name=aba_sintese, startrow=linha_resumo, index=False)
        if not resumo_mensalidades.empty:
            resumo_mensalidades.to_excel(writer, sheet_name=aba_sintese, startrow=linha_mensal, index=False)
        if not resumo_demais.empty:
            resumo_demais.to_excel(writer, sheet_name=aba_sintese, startrow=linha_demais, index=False)

        workbook = writer.book
        worksheet = workbook[aba_sintese]

        worksheet.cell(row=linha_resumo, column=1,
                       value="1. DADOS GERAIS DO CONTRATO").font = Font(bold=True, size=12)
        worksheet.cell(row=linha_mensal, column=1,
                       value="2. VALORES POR PRODUTO: MENSALIDADES").font = Font(bold=True, size=12)
        worksheet.cell(row=linha_demais, column=1,
                       value="3. VALORES POR PRODUTO: DEMAIS PRODUTOS").font = Font(bold=True, size=12)
        worksheet.cell(row=linha_avulsos, column=1,
                       value="4. HISTÓRICO DE COBRANÇAS AVULSAS").font = Font(bold=True, size=12)

        if avulsos_detalhado.empty:
            worksheet.cell(row=linha_avulsos + 1, column=1,
                           value="Sem cobranças avulsas no período selecionado.")
        else:
            avulsos_detalhado.to_excel(writer, sheet_name=aba_sintese, startrow=linha_avulsos, index=False)

        # 2. ABAS SEPARADAS
        df_resumo_geral.to_excel(writer, sheet_name='Resumo Geral', index=False)
        resumo_produtos_mensalidades.to_excel(writer, sheet_name='Prod. Mensalidades', index=False)
        resumo_produtos_demais.to_excel(writer, sheet_name='Demais Produtos', index=False)
        avulsos_detalhado.to_excel(writer, sheet_name='Cobranças Avulsas', index=False)
        df_fat.to_excel(writer, sheet_name='Base Histórica Completa', index=False)

    buf.seek(0)
    return buf


# ==========================================
# HELPER — Tabela HTML para impressão
# ==========================================
def _df_to_print_html(df, col_labels=None):
    """
    Converte um DataFrame em uma tabela HTML envolta em <div class="print-only-table">.
    Na tela fica escondida (display:none); na impressão aparece (display:block).
    col_labels: dict com {nome_coluna_original: "Rótulo Amigável"}.
    """
    if col_labels is None:
        col_labels = {c: c for c in df.columns}

    cols = [c for c in col_labels if c in df.columns]
    headers = "".join(f"<th>{col_labels[c]}</th>" for c in cols)

    rows = []
    for _, row in df.iterrows():
        cells = ""
        for c in cols:
            val = row[c]
            if isinstance(val, float):
                cells += f"<td style='text-align:right'>{val:,.2f}</td>".replace(",", "X").replace(".", ",").replace("X", ".")
            elif isinstance(val, (int,)):
                cells += f"<td style='text-align:right'>{val}</td>"
            else:
                cells += f"<td>{val}</td>"
        rows.append(f"<tr>{cells}</tr>")

    return (
        '<div class="print-only-table">'
        '<table style="width: 100% !important; max-width: 100% !important; table-layout: fixed !important;">'
        '<colgroup>'
        '<col style="width: 40%;">'
        '<col style="width: 16%;">'
        '<col style="width: 12%;">'
        '<col style="width: 12%;">'
        '<col style="width: 12%;">'
        '<col style="width: 8%;">'
        '</colgroup>'
        f'<thead><tr>{headers}</tr></thead>'
        f'<tbody>{"".join(rows)}</tbody></table></div>'
    )


# ==========================================
# INTERFACE — SIDEBAR
# ==========================================
with st.sidebar:
    st.markdown("### ⚙️ Parâmetros da Consulta")
    st.caption("Informe os códigos das empresas (um por linha) e clique em **Gerar Análise**.")

    empresas_texto = st.text_area(
        "Código(s) da Empresa Trabalho",
        placeholder="Ex:\n12345\n67890\n11223",
        height=120,
        help="Digite um código de empresa por linha. Pode consultar várias de uma vez.",
    )

    meses = st.number_input(
        "Quantidade de Meses",
        min_value=1,
        max_value=36,
        value=6,
        step=1,
        help="Quantos meses de faturamento deseja analisar.",
    )

    st.markdown("---")
    gerar = st.button("🚀  Gerar Análise de Saúde Financeira", use_container_width=True)


# ==========================================
# INTERFACE — ÁREA PRINCIPAL
# ==========================================
st.markdown(
    '<div class="hero-banner">'
    '<h1>📊 Saúde Financeira do Contrato</h1>'
    '<p>Análise integrada de faturamento, mensalidades e cobranças avulsas via SOC WebService</p>'
    '</div>',
    unsafe_allow_html=True,
)


def _parse_codigos_empresa(texto):
    """Extrai códigos numéricos válidos do text_area (um por linha)."""
    codigos = []
    for linha in texto.strip().splitlines():
        linha = linha.strip()
        if linha.isdigit() and int(linha) > 0:
            codigos.append(linha)
    return codigos


# ----- estado de execução -----
if gerar:
    codigos = _parse_codigos_empresa(empresas_texto) if empresas_texto.strip() else []

    if not codigos:
        st.warning("⚠️ Informe ao menos um **Código de Empresa Trabalho** válido na barra lateral.")
        st.stop()

    # Conectar ao SOC (cacheado)
    with st.spinner("🔌 Conectando ao WebService do SOC..."):
        try:
            cliente = conectar_soc()
        except Exception as e:
            st.error(f"Erro ao conectar no SOC: {e}")
            st.stop()

    resultados = []
    total_empresas = len(codigos)

    for idx, empresa_cod in enumerate(codigos, 1):
        label_empresa = f"Empresa {empresa_cod} ({idx}/{total_empresas})"
        with st.status(f"📡 {label_empresa} — Buscando dados...", expanded=True) as status_container:
            log = st.container()

            log.write(f"**Etapa 1/3** — [{empresa_cod}] Buscando funcionários ativos...")
            qtd_ativos = buscar_funcionarios_ativos(cliente, empresa_cod, log)

            log.write(f"**Etapa 2/3** — [{empresa_cod}] Buscando faturamento histórico (multithreading)...")
            df_faturamento = buscar_faturamento_historico(cliente, empresa_cod, meses, log)

            log.write(f"**Etapa 3/3** — [{empresa_cod}] Buscando regras de contrato...")
            df_precos_contrato = buscar_precos_contrato(cliente, empresa_cod, log)

            data_coleta = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            resultado = processar_dados(df_faturamento, df_precos_contrato, qtd_ativos, meses)

            if resultado is not None:
                resultado["empresa_codigo"] = empresa_cod
                resultado["idx"] = idx
                resultado["total_empresas"] = total_empresas
                resultado["data_coleta"] = data_coleta
                resultados.append(resultado)
                status_container.update(
                    label=f"✅ {label_empresa} — Concluído! — {data_coleta}",
                    state="complete", expanded=False,
                )
            else:
                status_container.update(
                    label=f"⚠️ {label_empresa} — Sem dados de faturamento",
                    state="error", expanded=False,
                )

    if not resultados:
        st.warning("⚠️ Nenhuma empresa retornou dados de faturamento no período informado.")
        st.stop()

    # Guardar no session_state para persistir entre reruns
    st.session_state["resultados"] = resultados


def _exibir_empresa(res):
    """Renderiza o dashboard de uma única empresa."""
    
    # Cabeçalho de Status para Impressão
    st.markdown(
        f'<div class="print-only-status">✅ Empresa {res.get("empresa_codigo", "?")} '
        f'({res.get("idx", "?")}/{res.get("total_empresas", "?")}) '
        f'— Concluído! — {res.get("data_coleta", "?")}</div>',
        unsafe_allow_html=True
    )

    # ── SEÇÃO 1: RESUMO GERAL (Metrics) ──
    st.markdown('<div class="section-title">📋 Resumo Geral do Contrato</div>', unsafe_allow_html=True)

    # Linha 1: Focada no nome da empresa (Razão Social costuma ser longa)
    c1, c2 = st.columns([2, 1])
    c1.metric("Razão Social", res["razao_social"])
    c2.metric("Funcionários Ativos", f'{res["qtd_ativos"]:,}'.replace(",", "."))

    # Linha 2: Focada no Faturamento
    c3, c4 = st.columns(2)
    c3.metric("Faturamento Total (R$)", f'R$ {res["faturamento_total"]:,.2f}'.replace(",", "X").replace(".", ",").replace("X", "."))
    c4.metric("Custo Mensalidade / Func.", f'R$ {res["custo_mensalidade_func"]:,.2f}'.replace(",", "X").replace(".", ",").replace("X", "."))

    # Linha 3: Focada nas datas
    c5, c6, c7 = st.columns(3)
    c5.metric("Período Avaliado", res["periodo"])
    c6.metric("Meses Analisados", res["meses"])
    c7.metric("Data Assinatura Contrato", str(res["data_assinatura"]))

    # ── SEÇÃO 2: TABELAS DE PRODUTOS ──
    col_config = {
        "NOME_PRODUTO": st.column_config.TextColumn("Produto", width="large"),
        "Total_Cobrado_Periodo": st.column_config.NumberColumn("Total Cobrado (R$)", format="R$ %.2f"),
        "Media_Vidas_Cobradas": st.column_config.NumberColumn("Vidas (Atual)", format="%.0f"),
        "Custo_Medio_Por_Vida": st.column_config.NumberColumn("Custo / Vida (Atual)", format="R$ %.2f"),
        "Faturamento_Minimo_Valor (R$)": st.column_config.NumberColumn("Fat. Mínimo (R$)", format="R$ %.2f"),
        "Faturamento_Minimo_Vidas": st.column_config.NumberColumn("Mín. Vidas", format="%.0f"),
    }
    
    html_cols = {
        "NOME_PRODUTO": "Produto",
        "Total_Cobrado_Periodo": "Total Cobrado (R$)",
        "Media_Vidas_Cobradas": "Vidas (Atual)",
        "Custo_Medio_Por_Vida": "Custo / Vida (Atual)",
        "Faturamento_Minimo_Valor (R$)": "Faturamento Mínimo (R$)",
        "Faturamento_Minimo_Vidas": "Mínimo Vidas"
    }

    # -- TABELA DE MENSALIDADES --
    st.markdown('<div class="section-title">📦 Valores por Produto: Mensalidades</div>', unsafe_allow_html=True)
    df_mens = res["resumo_mensalidades"].copy()
    if not df_mens.empty:
        for col in df_mens.select_dtypes(include='number').columns:
            df_mens[col] = df_mens[col].round(2)
        st.dataframe(df_mens, use_container_width=True, hide_index=True, column_config=col_config)
        st.markdown(_df_to_print_html(df_mens, html_cols), unsafe_allow_html=True)
    else:
        st.info("Nenhuma mensalidade encontrada no período.")

    # -- TABELA DE DEMAIS PRODUTOS --
    st.markdown('<div class="section-title">📦 Valores por Produto: Demais Produtos</div>', unsafe_allow_html=True)
    df_demais = res["resumo_demais"].copy()
    if not df_demais.empty:
        for col in df_demais.select_dtypes(include='number').columns:
            df_demais[col] = df_demais[col].round(2)
        st.dataframe(df_demais, use_container_width=True, hide_index=True, column_config=col_config)
        st.markdown(_df_to_print_html(df_demais, html_cols), unsafe_allow_html=True)
    else:
        st.info("Nenhum outro produto cobrado no período.")

    # ── SEÇÃO 3: COBRANÇAS AVULSAS ──
    st.markdown('<div class="section-title">🧾 Cobranças Avulsas</div>', unsafe_allow_html=True)

    if res["avulsos_detalhado"].empty:
        st.info("ℹ️ Sem cobranças avulsas no período selecionado.")
        st.markdown(
            '<div class="print-info" style="display:none;">Sem cobranças avulsas no período selecionado.</div>',
            unsafe_allow_html=True,
        )
    else:
        avulsos_sem_vidas = res["avulsos_detalhado"].drop(
            columns=["QUANTIDADE_VIDAS_ATIVAS"], errors="ignore"
        )
        st.dataframe(
            avulsos_sem_vidas,
            use_container_width=True,
            hide_index=True,
            column_config={
                "DATA_COBRANCA": st.column_config.TextColumn("Data Cobrança"),
                "NOME_PRODUTO": st.column_config.TextColumn("Produto", width="large"),
                "VALOR_TOTAL_PRODUTO": st.column_config.NumberColumn("Valor (R$)", format="R$ %.2f"),
            },
        )

        # Tabela HTML para impressão
        st.markdown(
            _df_to_print_html(avulsos_sem_vidas, {
                "DATA_COBRANCA": "Data Cobrança",
                "NOME_PRODUTO": "Produto",
                "VALOR_TOTAL_PRODUTO": "Valor (R$)",
            }),
            unsafe_allow_html=True,
        )


# ----- exibição dos resultados -----
if "resultados" in st.session_state:
    resultados = st.session_state["resultados"]

    if len(resultados) == 1:
        # Uma única empresa — exibe direto
        _exibir_empresa(resultados[0])
    else:
        # Múltiplas empresas — abas
        nomes_abas = [
            f"{r.get('empresa_codigo', '?')} — {r['razao_social'][:30]}"
            for r in resultados
        ]
        tabs = st.tabs(nomes_abas)
        for tab, res in zip(tabs, resultados):
            with tab:
                _exibir_empresa(res)

    # ── SEÇÃO 4: DOWNLOAD + IMPRIMIR ──
    st.markdown("---")

    # Gera um único Excel com todas as empresas (cada empresa = prefixo no nome da aba)
    if len(resultados) == 1:
        excel_buf = gerar_excel_em_memoria(resultados[0])
        nome_arquivo = f"Saude_Financeira_{resultados[0]['razao_social'].replace(' ', '_').replace('/', '')}.xlsx"
    else:
        # Múltiplas empresas em um único arquivo
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            for res in resultados:
                cod = res.get("empresa_codigo", "?")
                prefixo = cod[-5:]  # Últimos 5 dígitos para manter nome curto
                res["df_resumo_geral"].to_excel(writer, sheet_name=f'{prefixo}_Resumo', index=False)
                res["resumo_produtos"].to_excel(writer, sheet_name=f'{prefixo}_Produtos', index=False)
                avulsos = res["avulsos_detalhado"].drop(columns=["QUANTIDADE_VIDAS_ATIVAS"], errors="ignore")
                avulsos.to_excel(writer, sheet_name=f'{prefixo}_Avulsos', index=False)
                res["df_fat"].to_excel(writer, sheet_name=f'{prefixo}_Base', index=False)
        buf.seek(0)
        excel_buf = buf
        nome_arquivo = f"Saude_Financeira_{len(resultados)}_empresas.xlsx"

    col_dl, col_print, _ = st.columns([1, 1, 2])
    with col_dl:
        st.download_button(
            label="⬇️  Baixar Planilha Excel",
            data=excel_buf,
            file_name=nome_arquivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with col_print:
        import streamlit.components.v1 as components
        components.html(
            """
            <button onclick="window.parent.print()"
                    style="
                        width: 100%;
                        padding: 0.6rem 1rem;
                        background: linear-gradient(135deg, #005850, #007b6b);
                        color: white;
                        border: none;
                        border-radius: 10px;
                        font-weight: 600;
                        font-size: 0.92rem;
                        cursor: pointer;
                        letter-spacing: .02em;
                        transition: all .2s ease;
                        font-family: 'Inter', sans-serif;
                    "
                    onmouseover="this.style.transform='translateY(-2px)';this.style.boxShadow='0 6px 20px rgba(0,88,80,.4)'"
                    onmouseout="this.style.transform='';this.style.boxShadow=''"
            >🖨️&nbsp; Imprimir Relatório</button>
            """,
            height=50,
        )

else:
    # Estado inicial — sem dados
    st.markdown(
        """
        <div style="text-align:center; padding: 4rem 2rem; color: #64748b;">
            <p style="font-size:3rem; margin-bottom:.5rem;">🏦</p>
            <p style="font-size:1.1rem; font-weight:500;">
                Preencha os parâmetros na barra lateral e clique em
                <strong>Gerar Análise</strong> para começar.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )
